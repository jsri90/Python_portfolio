""" THIS CODE AIMS AT READING UPDATED DATA ROM EXCEL BACK INTO MATHCAD.
- PLEASE USE THIS FILE ONLY AFTER THE REQUIRED DATA HAS BEEN UPDATED IN SHEET 1 OF THE EXCEL SHEET.
- PLEASE MAKE SURE THE EXCEL FILE "Mathcad_EXcel.xlsx" IS SAVED AND CLOSED BEFORE THIS CODE IS EXECUTED """



# ****************************************** SECTION 0****************************************
""" THIS SECTION IMPORTS ALL THE REQUIRED PACKAGES AND MODULES """
import os
try:
    import re
except ImportError:
    print("Trying to install required module: re\n")
    os.system('python -m pip install re')
import re
try:
    import openpyxl
    from openpyxl.styles import Alignment
except ImportError:
    print("Trying to install required module: openpyxl\n")
    os.system('python -m pip install openpyxl')
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook



from shutil import copyfile

#*********************************** END SECTION 0 *********************************************************




#********************************* SECTION 1 **************************************************************

""" THIS SECTION OPENS THE WORKBOOK AND CREATES FILE HANDLES FOR EACH OF THE SHEETS """
wb = load_workbook(filename = 'Change_of_tags_Bypass1.xlsx')

Sht1 = wb["BKR"]
##col1 = Sht1.max_column # number of columns for sheet 1 data
##row1= Sht1.max_row


#***************************************************************************************
Sht2 = wb["REG"]
##col2= Sht2.max_column
##row2 = Sht2.max_row


#****************************************************************************
Sht3 = wb["APC_Meter"]
##col3= Sht3.max_column
##row3=Sht3.max_row


#******************************************************************************
Sht4 = wb["BKR-analogs"]
##col4= Sht3.max_column
##row4=Sht3.max_row
#*********************************************************************************
Sht5 = wb["Reg-Analogs"]
##row5 = Sht5.max_row
##col5 = Sht5.max_column


#****************************************************************************************
Sht6 = wb["APC-met"]
##row6=Sht6.max_row
##col6 = Sht6.max_column

###*********************************************************************************
Sht7 = wb["BCR"]
##row7=Sht7.max_row
##col7 = Sht7.max_column

#*********************************************************************************

Sht8 = wb["BinaryOutput-BKR"]
##row8 = Sht8.max_row
##col8 = Sht8. max_column


#****************************************************************************************
Sht9 = wb["BinaryOutput-Reg"]
##row9=Sht9.max_row
##col9 = Sht9.max_column

#*********************************************************************************
Sht10 = wb["LOCAL_TAGS"]
##row10=Sht10.max_row
##col10 = Sht10.max_column

#*********************************************************************************
Sht11 = wb["HMITags_Indirect"]
##row11 = Sht11.max_row
##col11 = Sht11.max_column





#******************************** END SECTION 1 ***********************************************************************

def replace_str(sh_r):
    max_r = sh_r.max_row
    max_c = sh_r.max_column
    s1 = "BKR200_"
    s2 = "BKR100_"
    s3 = "BKR3000_"
    s4 = "BKR4000_"
    s5 = "BKR1000_"
    s6 = "BKR2000_"
    s7 = "Reg3010"
    s8 = "Reg4010"
    s9 = "Reg1010"
    s10= "Reg2010"
    data1 = ""
    for i in range(1,max_c+1):
        for j in range(1, max_r+1):
##            sh_w.column_dimensions[get_column_letter(i)].width = 45
            data = str(sh_r.cell(row = j, column = i).value)
            if s1 in data:
                data1 = data.replace(s1,"BKRB1000_")
            elif s2 in data:
                data1= data.replace(s2,"BKRC1000_")
            elif s3 in data:
                data1= data.replace(s3,"BKRB1100_")
            elif s4 in data:
                data1= data.replace(s4, "BKRB1200_")
            elif s5 in data:
                data1= data.replace(s5, "BKRC1100_")
            elif s6 in data:
                data1 = data.replace(s6,"BKRC1200_")
            elif s7 in data:
                data1 = data.replace(s7,"RegB1110")
            elif s8 in data:
                data1 = data.replace(s8,"RegB1210")
            elif s9 in data:
                data1 = data.replace(s9,"RegC1110")
            elif s10 in data:
                data1= data.replace(s10, "RegC1210")

            sh_r.cell(row = j, column = i, value = data1)




#****************************************** SECTION 4*******************************************************************






            
##    k.text = str(Sht1.cell(row = row_x, column = column_y).value)

    
replace_str(Sht1) # gen

replace_str(Sht2)

replace_str(Sht4)

replace_str(Sht5)

replace_str(Sht8)

replace_str(Sht9)

##write_val(Sht5, Sht6, Sht7,row5, row6, max_col) # 2a
##
##write_val(Sht8, Sht9, Sht10,row8, row9, max_col) # 2b
##
##write_val(Sht11, Sht12, Sht13,row11, row12, max_col) # 2c
##
##write_val(Sht14, Sht15, Sht16,row14, row15, max_col) #4a
##
##write_val(Sht17, Sht18, Sht19,row17, row18, max_col) #4b
##
##write_val(Sht20, Sht21, Sht22,row20, row21, max_col) # ex
##
##write_val(Sht23, Sht24, Sht25,row23, row24, max_col) # msu
##


    
    


    
    
#*************************************** END SECTION 4*****************************************************************

wb.save("Change_of_tags_Bypass1.xlsx")













            
