



# ****************************************** SECTION 0****************************************
""" THIS SECTION IMPORTS ALL THE REQUIRED PACKAGES AND MODULES """
import os
try:
    import re
except ImportError:
    print("Trying to install required module: re\n")
    os.system('python -m pip install re')
import re
try:
    import openpyxl
    from openpyxl.styles import Alignment
except ImportError:
    print("Trying to install required module: openpyxl\n")
    os.system('python -m pip install openpyxl')
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook



from shutil import copyfile

#*********************************** END SECTION 0 *********************************************************

#tag_proc = wb.create_sheet("Tag processor")



#********************************* SECTION 1 **************************************************************

""" THIS SECTION OPENS THE WORKBOOK AND CREATES FILE HANDLES FOR EACH OF THE SHEETS """
wb = load_workbook(filename = 'Rosepine_Settings_Group.xlsx')

Sht1 = wb["Sheet1"]
max_col1 = Sht1.max_column # number of columns for sheet 1 data

max_row1 = Sht1.max_row

Sht2 = wb.create_sheet("Settings_Group")







#******************************** END SECTION 1 ***********************************************************************

##def write_col(sh_r,sh_w, rw_read, rw_write, cl):
##
##    for i in range(1,cl+1):
##        sh_w.column_dimensions[get_column_letter(i)].width = 45
##        data = sh_r.cell(row = rw_read, column = i).value
##        sh_w.cell(row = rw_write, column = i, value = data)

for i in range(2, max_row1+1):
    dev = int(Sht1.cell(row = i, column = 1).value)
    data_SG1 = str(Sht1.cell(row = i, column = 2).value)
    data_SG2 = str(Sht1.cell(row = i, column = 3).value)
    
    list_SG1 = data_SG1.split(',')
    listInt_SG1 = [int(i) for i in list_SG1]
    print(f'SG1 = {listInt_SG1}')
    
    list_SG2 = data_SG2.split(',')
    listInt_SG2 = [int(i) for i in list_SG2]
    print(f'SG2 = {listInt_SG2}')

    
    Sht2.cell(row = i, column = 1, value = dev)
    columnVal = 2
    
    for sg1 in listInt_SG1:
        Sht2.cell(row = i, column = columnVal, value = sg1)
        columnVal += 1
        Sht2.cell(row = i, column = columnVal, value = 1)
        columnVal += 1

    for sg2 in listInt_SG2:
        Sht2.cell(row = i, column = columnVal, value = sg2)
        columnVal += 1
        Sht2.cell(row = i, column = columnVal, value = 2)
        columnVal += 1

    while ( columnVal <= 13):
        Sht2.cell(row = i, column = columnVal, value = 0)
        columnVal += 1
        Sht2.cell(row = i, column = columnVal, value = 0)
        columnVal += 1
        
       
##                   
##    if data:
##        data1 = data.strip()
##        Sht2.cell(row = i, column = j, value = data1)




#****************************************** SECTION 4*******************************************************************







    
    


    
    
#*************************************** END SECTION 4*****************************************************************

wb.save('Rosepine_Settings_Group.xlsx')













            
