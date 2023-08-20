# importing all necessary modules from OpenPyxl
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from openpyxl.styles import colors # used for coloring cells
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font # Connect cell styles
from openpyxl.workbook import Workbook # used for creating a new workbook



# create a new workbook
wb = openpyxl.Workbook()


# syntax to create a new sheet in excel using openpyxl
tag_proc = wb.create_sheet("Tag processor")


# syntax to iterate through sheets in a workbook and get their indices:
wb = load_workbook(t) # load workbook to be modified
for sheet in wb.worksheets:
    #print(sheet.title)
    #print(wb.worksheets.index(sheet))
    if wb.worksheets.index(sheet) == 2:
        dds_modbus(sheet)


# change title of existing sheet :
wb["Sheet"].title = "Mathcad_Excel_raw"

# set the width and height of the columns and rows
Sht2.column_dimensions['A'].width = 45
sheet.row_dimensions[1].height = 70


# creates a handle for loading a workbook
wb = load_workbook(filename = 'Bunkie_manipulation.xlsx')

# save workbook
wb.save("RTAC settings- voltage- tag manipulation_1.xlsx")

# create file handle for loading a sheet and accessing the max rows and columns
Sht1 = wb["Bunkie- binaries"]
max_col1 = Sht1.max_column # number of columns for sheet 1 data
max_row1 = Sht1.max_row

max_row1 = len(Sht['C']) # get length of speific column in excel sheet


# reading value from a particular cell
bi = str(Sht1.cell(row = i, column = 1).value)

# writing value to a particular cell
Sht1.cell(row = i, column = 3, value = word_bi)


# updating pattern of font and color in a particulr cell
               CELL2 = Sht.cell(row = i+1, column = 2)
CELL1.fill = PatternFill(fill_type = 'solid',start_color = 'ffff00', end_color = 'ffff00')
CELL1.font = Font(color = 'ff0000')

# get index of specific sheet in a workbook
print(wb.worksheets.index(wb["Modbus Map"]))


# returns length of a column containing valid values
def count_rows_in_column(Sht, letter):
    count = 0
    blank_count = 0
    for k in Sht[letter]:
        if k != None:
            count += 1
        else:
            blank_count += 1

        if blank_count == 2:
            break

    return count








