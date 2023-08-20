# MPC - Summrall

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, cell
from openpyxl.styles import Alignment

import os
import xml.etree.ElementTree as ET

from shutil import copyfile, move

#******************getting saubstation mae from user************************************
fileName = 'Access_Points.xlsx'

# returns length of a column containing valid values
def count_rows_in_column(Sht, letter):
    r = 1
    c= cell.column_index_from_string(letter)
    count = 0
    blank_count = 0
    while(True):
        if (Sht.cell(row = r, column = c).value) != None:
            count += 1
        else:
            blank_count += 1
        r += 1

        if blank_count == 1:
            break

    return count

##file = 'Sample_Templates/XXXX_VRTU_DNP.xml'
##s1 = 'Test1.xml'
##copyfile(file,s1)
cwd = os.getcwd()
print(cwd)
# loading devices list
wb = load_workbook(fileName)
Sht1 = wb['Master_Device_List']







def ckt_dev(Sht):
    circuit_devices = {}

    num_circuits = count_rows_in_column(Sht, 'A')


    for ckt in range(2, num_circuits + 1):
        ap_num = str(Sht.cell(row = ckt, column = 1).value)
        tap_num = str(Sht.cell(row = ckt, column = 2).value)
        port_num = str(Sht.cell(row = ckt, column = 4).value)

        circuit_devices[ap_num] = [tap_num,port_num]


    return circuit_devices
    
    
DeviceCircuits = ckt_dev(Sht1)



#********************************change Srv connection parameters for SIM ********************************************************************
#loading sample SIM Srv

ap_sample = 'Sample_Templates/Sample_1.xml'
tap_sample = 'Sample_Templates/Sample_1_Transparent.xml'


# modifying parameters:

def change_ap_parameters(root, port_num, fname):

    for tags in root:
        if tags.text != None and '9999' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('9999', port_num)
            tags.text = temp

        if tags.text != None and 'Sample_1' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('Sample_1', fname)
            tags.text = temp


        change_ap_parameters(tags, port_num, fname)


def change_tap_parameters(root, src, dest, leg, fTap):

    for tags in root:
        if tags.text != None and 'XXXX' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('XXXX', src)
            tags.text = temp

        if tags.text != None and 'YYYY' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('YYYY', dest)
            tags.text = temp

        if tags.text != None and '111' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('111', leg)
            tags.text = temp

        if tags.text != None and 'Sample_1_Transparent' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('Sample_1_Transparent', fTap)
            tags.text = temp

        if tags.text != None and 'AAAA' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('AAAA', 'TRUE')
            tags.text = temp


        change_tap_parameters(tags, src, dest, leg, fTap)


# writing Access point parameters:
def write_ap_tags(tree, file, *args):

    if args[0] != None:
        port_number = args[0]
        #print('Recloser name: ', server_DNP)
    else:
        print('Enter a valid server DNP address')

    if args[1] != None:
        file_name = args[1]
        #print('Recloser name: ', server_DNP)
    else:
        print('Enter a valid file name')


    root1 = tree.getroot()
    change_ap_parameters(root1, port_number, file_name)
    file_string = ET.tostring(root1)

    with open(file, 'wb') as f:
        f.write(file_string)

# writing Transparent Access point parameters:
def write_tap_tags(tree, file, *args):

    if args[0] != None:
        source = args[0]
        #print('Recloser name: ', server_DNP)
    else:
        print('Enter a valid source')

    if args[1] != None:
        destination = args[1]
        #print('Recloser name: ', server_DNP)
    else:
        print('Enter a valid destination')

    if args[2] != None:
        legacy = args[2]
        #print('Recloser name: ', server_DNP)
    else:
        print('Enter a valid legacy cmd ')

    if args[3] != None:
        tap_fName = args[3]
        #print('Recloser name: ', server_DNP)
    else:
        print('Enter a valid tap file name ')



    root1 = tree.getroot()
    change_tap_parameters(root1, source, destination, legacy, tap_fName)
    file_string = ET.tostring(root1)

    with open(file, 'wb') as f:
        f.write(file_string)
        

num_devices = count_rows_in_column(Sht1, 'A')

legacy_cmd = 1
for ap in range(2, num_devices + 1):

    ap_name = str(Sht1.cell(row = ap, column = 1).value)
    tap_name = str(Sht1.cell(row = ap, column = 2).value)
    location = str(Sht1.cell(row = ap, column = 3).value)
    ap_port = str(Sht1.cell(row = ap, column = 4).value)
    print('AP port: ', ap_port)

    
    ap_file = str(ap_name + '_AP.xml')
    print('filename :', ap_file)
    copyfile(ap_sample, ap_file)

    tap_file = str(tap_name + '_Transparent.xml')
    tap_file1 = str(tap_name + '_Transparent')
    tap_fileName = tap_file1  # for file name inside the xml file
    print('filename :', tap_file)
    copyfile(tap_sample, tap_file)

    #modifying paramters
    tree_ap = ET.parse(ap_file)
    root_ap = tree_ap.getroot()

    tree_tap = ET.parse(tap_file)
    root_tap = tree_tap.getroot()
    
    write_ap_tags(tree_ap, ap_file, ap_port, ap_name)
    write_tap_tags(tree_tap, tap_file, ap_name, tap_name, str(legacy_cmd), tap_fileName)
    
    sim_srv_dir = os.path.join(cwd, str('AP and TAP/' + location))

    if not os.path.exists(sim_srv_dir):
        os.makedirs(sim_srv_dir)

    loc1 = cwd + '\\' + ap_file
    loc2 = sim_srv_dir + '\\' + ap_file
    move(loc1, loc2)

    loc1 = cwd + '\\' + tap_file
    loc2 = sim_srv_dir + '\\' + tap_file
    move(loc1, loc2)
    legacy_cmd += 1



#**************************************************************************************************



    


    

wb.save(fileName)





    
    



    
            
            







