# MPC - Summrall

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, cell
from openpyxl.styles import Alignment

import os
import xml.etree.ElementTree as ET

from shutil import copyfile, move

#******************getting saubstation mae from user************************************
file_name = input('Enter file name of Excel spreadsheet: ')

# returns length of a column containing valid values
def count_rows_in_column(Sht, letter):
    r = 1
    c= cell.column_index_from_string(letter)
    count = 0
    blank_count = 0
    while(True):
        if (Sht.cell(row = r, column = c).value) != None:
            count += 1
        else:
            blank_count += 1
        r += 1

        if blank_count == 1:
            break

    return count

##file = 'Sample_Templates/XXXX_VRTU_DNP.xml'
##s1 = 'Test1.xml'
##copyfile(file,s1)
cwd = os.getcwd()
print(cwd)
# loading devices list
wb = load_workbook(file_name)
Sht1 = wb['Sheet1']
Fdr = wb['Feeders']
FldDev = wb['Field_devices_ mapping']
SG = wb['Field_Device_SG']
Fld_AutoMap = wb['Field_Device_AutoMap']
Flt_Tags = wb['Fault_Tags']
Gateway = wb['Gateway_Printing']

Xfmr_sheet = wb['Prg_XFMR']






def ckt_dev(Sht):
    circuit_devices = {}

    num_circuits = count_rows_in_column(Sht, 'A')


    for ckt in range(2, num_circuits + 1):
        ckt_num = str(Sht.cell(row = ckt, column = 1).value)
        rec_num = str(Sht.cell(row = ckt, column = 5).value)

        if ckt_num not in circuit_devices:
            circuit_devices[ckt_num] = [rec_num]
        else:
            circuit_devices[ckt_num].append(rec_num)

    return circuit_devices
    
    
DeviceCircuits = ckt_dev(Sht1)


#***************************************** 651R DNP ****************************************************

# creating client device connections - Field devices
#loadng sample file creating recloser files



#Changing device names in fault data programs:
def fault_prog(root, dev_name,temp_tag):

    for tags in root:
        if tags.text != None and temp_tag in tags.text:
            temp = str(tags.text)
            temp = temp.replace(temp_tag, dev_name)
            tags.text = temp

        fault_prog(tags,dev_name, temp_tag)


#Changing device names in fault data programs:
def fault_title(root, fdr_name,temp_tag):

    for tags in root:
        if tags.text != None and temp_tag in tags.text:
            temp = str(tags.text)
            temp = temp.replace(temp_tag, fdr_name)
            tags.text = temp

        fault_title(tags,fdr_name, temp_tag)


        



# writing the modified tags into XML
def write_tags(tree, file, *args):
    if args[0] != None:
        feeder_name = str(args[0])
        #print('Recloser name: ', client_device)
    else:
        print('Enter a valid feeder name')

    if args[1] != None:
        feeder_temp_tag = str(args[1])
        #print('IP address : ', IP_adr)
    else:
        print('Enter a valid feeder temp tag')

    if args[2] != None:
        device_name = str(args[2])
        #print('Server DNP address: ', server_DNP)
    else:
        print('Enter a valid device name')


    if args[3] != None:
        device_temp_tag = str(args[3])
        #print('Server IP port :', server_IP)
    else:
        print('Enter a valid device temp tag')
    
    root1 = tree.getroot()
    fault_prog(root1, device_name, device_temp_tag)
    fault_title(root1, feeder_name, feeder_temp_tag)
    file_string = ET.tostring(tree.getroot())

    with open(file, 'wb') as f:
        f.write(file_string)

#-----------------------------------
fault_sample = 'Sample_Templates/Circuit_ZZZZ_Fault_Data.xml'

for circuit in DeviceCircuits:

    devices_list = DeviceCircuits[circuit]


    str_file = str('Circuit_' + str(circuit) + '_Fault_Data.xml')
    copyfile(fault_sample, str_file)
    #modifying parameters:
    tree = ET.parse(str_file)
    root = tree.getroot()
    c1 = 'A'

    for dev_num in devices_list:
        s1 = c1 * 3
        write_tags(tree, str_file, circuit, 'ZZZZ', dev_num, s1)
        c1 = chr(ord(c1) + 1)


    fault_dir = os.path.join(cwd, 'DAC/Fault Data')


    if not os.path.exists(fault_dir):
        os.makedirs(fault_dir)


    loc1 = cwd + '\\' +str_file
    loc2 = fault_dir + '\\' + str_file
    move(loc1, loc2)
#************************************************************************************************************

#**************************************** Feeders***********************************************************
DeviceDefinition = 'FdrDef_Standard'
FdrCapacity = str(int(795 * 0.9))
SourceGrp = '1'
FieldFirstBi = '0'
FieldFirstAi = '0'
FieldFirstBo = '0'
FieldFirstAo = '0'
FieldFirstCnt = '0'

FirstBi = 'Armed'
FirstAi = '0'
FirstBo = 'Exclude_CMD'
FirstAo = '0'
FirstCnt = '0'

i = 2
for circuit in DeviceCircuits:
    Fdr.cell(row = i, column = 1, value = str(circuit)) # Feeder Num
    Fdr.cell(row = i, column = 3, value = str(DeviceDefinition)) # Device Definition
    Fdr.cell(row = i, column = 4, value = str(FdrCapacity)) # Feeder Capacity
    Fdr.cell(row = i, column = 5, value = str(FdrCapacity)) # Feeder Temp Capacity
    Fdr.cell(row = i, column = 6, value = str(SourceGrp)) # Source Group
    Fdr.cell(row = i, column = 7, value = str(FieldFirstBi)) # Field First Bi
    Fdr.cell(row = i, column = 8, value = str(FieldFirstAi)) # Field First Ai
    Fdr.cell(row = i, column = 9, value = str(FieldFirstBo)) # Field First Bo
    Fdr.cell(row = i, column = 10, value = str(FieldFirstAo)) # Field Fist Ao
    Fdr.cell(row = i, column = 11, value = str(FieldFirstCnt)) # Field First Cnt
    Fdr.cell(row = i, column = 12, value = str(circuit)) # SCADA Link
    Fdr.cell(row = i, column = 13, value = str(FirstBi)) # First Bi
    Fdr.cell(row = i, column = 14, value = str(FirstAi)) # First Ai
    Fdr.cell(row = i, column = 15, value = str(FirstBo)) # First Bo
    Fdr.cell(row = i, column = 16, value = str(FirstAo)) # First Ao
    Fdr.cell(row = i, column = 17, value = str(FirstCnt)) # First Cnt

    i += 1
    
    
#************************************************************************************************************

#*********************************** Field Devices*************************************************************
DevDef1 = 'RecDef_Standard'
DevDef2 = 'RecDef_Standard_withCommFail'

FieldFirstBi = 'Recloser_Status'
FieldFirstAi = 'IA'
FieldFirstBo = 'SS1'
FieldFirstAo = '0'
FieldFirstCnt = '0'


FirstBi = 'Recloser_Status'
FirstAi = 'IA'
FirstBo = 'SS1'
FirstAo = '0'
FirstCnt = '0'

num_devices = count_rows_in_column(Sht1, 'E')

for d in range(2, num_devices + 1):
    DeviceNumber = str(Sht1.cell(row = d, column = 5).value)
    NormalStatus = str(Sht1.cell(row = d, column = 6).value)
    DeviceType = str(Sht1.cell(row = d, column = 7).value)
    DeviceCapacity = str(Sht1.cell(row = d, column = 8).value)

    FldDev.cell(row = d, column = 1, value = str(DeviceNumber)) # Device Number

    if DeviceType == 'Standard':
        FldDev.cell(row = d, column = 3, value = str(DevDef1)) # Device Type
    else:
        FldDev.cell(row = d, column = 3, value = str(DevDef2)) # Device Type


    if NormalStatus == 'NC':
        FldDev.cell(row = d, column = 4, value = 'FALSE') # Normal Status
    else:
        FldDev.cell(row = d, column = 4, value = 'TRUE') # Normal Status

    FldDev.cell(row = d, column = 5, value = str(DeviceCapacity)) # Device Capacity
    FldDev.cell(row = d, column = 6, value = str(DeviceCapacity)) # Device Temp Capacity


    FldDev.cell(row = d, column = 7, value = 'Set2') # Side A PT
    FldDev.cell(row = d, column = 8, value = 'Set1') # Side B PT


    FldDev.cell(row = d, column = 9, value = str(FieldFirstBi)) # Field First Bi
    FldDev.cell(row = d, column = 10, value = str(FieldFirstAi)) # Field First Ai
    FldDev.cell(row = d, column = 11, value = str(FieldFirstBo)) # Field First Bo
    FldDev.cell(row = d, column = 12, value = str(FieldFirstAo)) # Field First Ao
    FldDev.cell(row = d, column = 13, value = str(FieldFirstCnt)) # Field First Cnt

    FldDev.cell(row = d, column = 17, value = str(DeviceNumber)) # Device Number
    
    FldDev.cell(row = d, column = 18, value = str(FirstBi)) # First Bi
    FldDev.cell(row = d, column = 19, value = str(FirstAi)) # First Ai
    FldDev.cell(row = d, column = 20, value = str(FirstBo)) # First Bo
    FldDev.cell(row = d, column = 21, value = str(FirstAo)) # First Ao
    FldDev.cell(row = d, column = 22, value = str(FirstCnt)) # First Cnt

#**********************************************************************************************************************

#********************************************** Settings Group***************************************************************
##
##for d in range(2, num_devices + 1):
##    dev = int(Sht1.cell(row = d, column = 5).value)
##    data_SG1 = str(Sht1.cell(row = d, column = 10).value)
##    data_SG2 = str(Sht1.cell(row = d, column = 11).value)
##    
##    list_SG1 = data_SG1.split(',')
##    listInt_SG1 = [int(i) for i in list_SG1]
##    print(f'SG1 = {listInt_SG1}')
##    
##    list_SG2 = data_SG2.split(',')
##    listInt_SG2 = [int(i) for i in list_SG2]
##    print(f'SG2 = {listInt_SG2}')
##
##    
##    SG.cell(row = d, column = 1, value = dev)
##    columnVal = 2
##    
##    for sg1 in listInt_SG1:
##        SG.cell(row = d, column = columnVal, value = sg1)
##        columnVal += 1
##        SG.cell(row = d, column = columnVal, value = 1)
##        columnVal += 1
##
##    for sg2 in listInt_SG2:
##        SG.cell(row = d, column = columnVal, value = sg2)
##        columnVal += 1
##        SG.cell(row = d, column = columnVal, value = 2)
##        columnVal += 1
##
##    while (columnVal <= 13):
##        SG.cell(row = d, column = columnVal, value = 0)
##        columnVal += 1
##        SG.cell(row = d, column = columnVal, value = 0)
##        columnVal += 1
##
##
##    SG.cell(row = d, column = 14, value = 'REC')


#******************************** Field Device AutoMap****************************************************************************    

FirstBi = 'Reloser_Status'
FirstAi = 'IA'
FirstBo = 'SS1'
FirstAo = '0'
FirstCnt = '0'

NumBi = '43'
NumAi = '18'
NumBo = '34'
NumAo = '0'
NumCnt = '0'



for d in range(2, num_devices + 1):
    dev = int(Sht1.cell(row = d, column = 5).value)

    Fld_AutoMap.cell(row = d, column = 1, value = dev)

    Fld_AutoMap.cell(row = d, column = 2, value = str(FirstBi)) # First Bi
    Fld_AutoMap.cell(row = d, column = 3, value = str(FirstAi)) # First Ai
    Fld_AutoMap.cell(row = d, column = 4, value = str(FirstBo)) # First Bo
    Fld_AutoMap.cell(row = d, column = 5, value = str(FirstAo)) # First Ao
    Fld_AutoMap.cell(row = d, column = 6, value = str(FirstCnt)) # First Cnt

    Fld_AutoMap.cell(row = d, column = 7, value = str(NumBi)) # Num Bi
    Fld_AutoMap.cell(row = d, column = 8, value = str(NumAi)) # Num Ai
    Fld_AutoMap.cell(row = d, column = 9, value = str(NumBo)) # Num Bo
    Fld_AutoMap.cell(row = d, column = 10, value = str(NumAo)) # Num Ao
    Fld_AutoMap.cell(row = d, column = 11, value = str(NumCnt)) # Num Cnt

    


#********************* Fault Tags*****************************************************************************
FLT_Curr_Dest = ['AI_00018_IA', 'AI_00019_IB', 'AI_00020_IC', 'AI_00021_IG']
FLT_Curr_Src = ['IA', 'IB', 'IC', 'IG']
count = 0
DMSName = input('Enter name of SCADA map for Fault Tags : ')


Flt_Tags.column_dimensions[get_column_letter(1)].width = 55
Flt_Tags.column_dimensions[get_column_letter(3)].width = 55

shtRow = 1


for d in range(2, num_devices + 1):
    dev = str(Sht1.cell(row = d, column = 5).value)
    for f in range(len(FLT_Curr_Dest)):
        Fault_String_Dest = str(DMSName + '_DNP.SEL_651R2_' + str(dev) + '_' + FLT_Curr_Dest[f] + '_FAULT_CURRENT')
        Flt_Tags.cell(row = shtRow, column = 1, value = str(Fault_String_Dest))

        Fault_String_Src = str('Fault_Data.SEL_651R2_' + str(dev) + '_' + FLT_Curr_Src[f] + '_FAULT_CURRENT')
        Flt_Tags.cell(row = shtRow, column = 3, value = str(Fault_String_Src))

        shtRow += 1
        
        count += 1



#**************************************** ****************************************************************************

#*****************************Gateway printing***********************************************************************



def gate_print(device, port, IP):



    SimDNP_Port = str(port)

    data1 = 'GatewayConfigurator.m_AddRoutingRule(	Simulator_DNP_Port:=\''
    data2 = SimDNP_Port
    data3 = '\', Equipment_Ip:=\''
    data4 = IP
    data5 = '\', Equipment_DNP_Port:=\'20000\'); //'
    data7 = str('R'+ str(device))

    data_final = str(data1 + data2 + data3 + data4 + data5 + ' ' + data7)
    return data_final

        

    #print('**************************************************************** \n\n')


Gateway.column_dimensions[get_column_letter(1)].width = 150

port = 20000
last_octet = 0
rowNum = 1

for d in range(2, num_devices + 1):
    dev = str(Sht1.cell(row = d, column = 5).value)
    IP_address = str(Sht1.cell(row = d, column = 14).value)
    port1 = str(port)
    cell_value = gate_print(dev, port1, IP_address)
    Gateway.cell(row = rowNum, column = 1, value = cell_value)
    port += 1
    rowNum += 1

    
    
#****************************** DMS MAPS*********************************************************

DMS_bi = wb['DMS_MAP-BI']
DMS_bo = wb['DMS_MAP-BO']
DMS_ai = wb['DMS_MAP-AI']


feeder_bi_len = count_rows_in_column(DMS_bi, 'A') # length of reference Feeder template tags going to SCADA
recloser_bi_len = count_rows_in_column(DMS_bi, 'D')# length of reference recloser template tags going to SCADA
    

feeder_bo_len = count_rows_in_column(DMS_bo, 'A')# length of reference Feeder template tags going to SCADA
recloser_bo_len = count_rows_in_column(DMS_bo, 'D')# length of reference recloser template tags going to SCADA    


recloser_ai_len = count_rows_in_column(DMS_ai, 'A') # length of reference recloser template tags going to SCADA
    

DMS_bi.column_dimensions['B'].width = 75
DMS_bi.column_dimensions['E'].width = 75


DMS_bo.column_dimensions['B'].width = 75
DMS_bo.column_dimensions['E'].width = 75


DMS_ai.column_dimensions['B'].width = 75

    
# modifying SCADA feeder bi, bo and ai tags
feeder_bi_index = 2
feeder_bo_index = 2

for circuit in DeviceCircuits:
    
    # modifying bi tags
    for k in range(2, feeder_bi_len + 1):
        feeder_bi_tag = str(DMS_bi.cell(row = k, column = 1).value)
        feeder_bi_tag = feeder_bi_tag.replace('YYYY', str(circuit))
        DMS_bi.cell(row = feeder_bi_index, column = 2, value = feeder_bi_tag)
        feeder_bi_index += 1
        
    # modifying bo tags
    for m in range(2, feeder_bo_len + 1):
        feeder_bo_tag = str(DMS_bo.cell(row = m, column = 1).value)
        feeder_bo_tag = feeder_bo_tag.replace('YYYY', str(circuit))
        DMS_bo.cell(row = feeder_bo_index, column = 2, value = feeder_bo_tag)
        feeder_bo_index += 1


# modifying SCADA recloser bi and bo tags. No ai tags for reclosers
recloser_bi_index = 2
recloser_bo_index = 2
recloser_ai_index = 2
for j in range(2, num_devices + 1):
    dev = str(Sht1.cell(row = j, column = 5).value)
    
    # modifying bi tags
    for k in range(2, recloser_bi_len + 1):
        recloser_bi_tag = str(DMS_bi.cell(row = k, column = 4).value)
        recloser_bi_tag = recloser_bi_tag.replace('XXXX', dev)
        DMS_bi.cell(row = recloser_bi_index, column = 5, value = recloser_bi_tag)
        recloser_bi_index += 1
        
    # modifying bo tags
    for m in range(2, recloser_bo_len + 1):
        recloser_bo_tag = str(DMS_bo.cell(row = m, column = 4).value)
        recloser_bo_tag = recloser_bo_tag.replace('XXXX', dev)
        DMS_bo.cell(row = recloser_bo_index, column = 5, value = recloser_bo_tag)
        recloser_bo_index += 1

    # modifying ai tags
    for n in range(2, recloser_ai_len + 1):
        recloser_ai_tag = str(DMS_ai.cell(row = n, column = 1).value)
        recloser_ai_tag = recloser_ai_tag.replace('XXXX', dev)
        DMS_ai.cell(row = recloser_ai_index, column = 2, value = recloser_ai_tag)
        recloser_ai_index += 1

    


#********************************change Srv connection parameters for SIM ********************************************************************
#loading sample SIM Srv

sim_srv_sample = 'Sample_Templates/Srv_SEL_651R2_XXXX_DNP.xml'


# modifying parameters:

def change_sim_srv_parameters(root, server_DNP, map_name, sim_ip_port):

    for tags in root:
        if tags.text != None and '65519' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('65519', server_DNP)
            tags.text = temp

        if tags.text != None and 'XXXX' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('XXXX', map_name)
            tags.text = temp


        if tags.text != None and '65534' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('65534', sim_ip_port)
            tags.text = temp

        change_sim_srv_parameters(tags, server_DNP, map_name, sim_ip_port)


# writing VRTU parameters:
def write_sim_srv_tags(tree, file, *args):

    if args[0] != None:
        server_DNP = args[0]
        #print('Recloser name: ', server_DNP)
    else:
        print('Enter a valid server DNP address')

    if args[1] != None:
        map_name = args[1]
        #print('Map name : ', map_name)
    else:
        print('Enter a valid map name')

    if args[2] != None:
        sim_IP_port = args[2]
    else:
        print('Enter a valid IP port')


    root1 = tree.getroot()
    change_sim_srv_parameters(root1, server_DNP, map_name, sim_IP_port)
    file_string = ET.tostring(root1)

    with open(file, 'wb') as f:
        f.write(file_string)

for sim_srv in range(2, num_devices + 1):
    device_name = str(Sht1.cell(row = sim_srv, column = 5).value)
    ip_port_sim = str(Sht1.cell(row = sim_srv, column = 12).value)
    print('SIM IP port: ', ip_port_sim)
    sim_srv_DNP_addr = str(Sht1.cell(row = sim_srv, column = 13).value)
    sim_srv_file = str('Srv_' + device_name + '_DNP.xml')

    copyfile(sim_srv_sample, sim_srv_file)

    #modifying paramters
    tree_sim_srv = ET.parse(sim_srv_file)
    root_sim_srv = tree_sim_srv.getroot()
    write_sim_srv_tags(tree_sim_srv, sim_srv_file, sim_srv_DNP_addr, device_name, ip_port_sim)
    sim_srv_dir = os.path.join(cwd, 'SIM/Field_Devices')

    if not os.path.exists(sim_srv_dir):
        os.makedirs(sim_srv_dir)

    loc1 = cwd + '\\' + sim_srv_file
    loc2 = sim_srv_dir + '\\' + sim_srv_file
    move(loc1, loc2)



#**************************************************************************************************



    
#**************************Server Map *******************************************************************
#loading sample server map
server_map_sample = 'Sample_Templates/SEL_651R2_XXXX_DNP.xml'


# modifying server map name
def change_server_map_name(root, server_map):

    for tags in root:
        if tags.text != None and 'XXXX' in tags.text:
            temp = str(tags.text)
            temp = temp.replace('XXXX',server_map)
            tags.text = temp

        change_server_map_name(tags, server_map)

# writing server map name

def write_server_map(tree, file, map_name):

    root1 = tree.getroot()
    change_server_map_name(root1, map_name)

    file_string = ET.tostring(tree.getroot())

    with open(file, 'wb') as f:
        f.write(file_string)


for srvMap in range(2, num_devices + 1):
    recloser_name = str(Sht1.cell(row = srvMap, column = 5).value)
    server_map_file = str('SEL_651R2_' + recloser_name + '_DNP.xml')

    copyfile(server_map_sample, server_map_file)

    #modifying parameters
    tree_server = ET.parse(server_map_file)
    root_server = tree_server.getroot()
    
    write_server_map(tree_server, server_map_file, recloser_name)
    server_map_dir = os.path.join(cwd, 'SIM/DNP Server Maps')

    if not os.path.exists(server_map_dir):
        os.makedirs(server_map_dir)

    loc1 = cwd + '\\' + server_map_file
    loc2 = server_map_dir +'\\' + server_map_file
    move(loc1, loc2)

#*********************************************************************************************************************************


###******************************** Program XFMR****************************************************************************
##
##XFMR_set = set()
##Xfmr_sheet.column_dimensions[get_column_letter(1)].width = 150
##
##
##
##
##for x1 in range(2, num_devices + 1):
##    XfmrName = str(Sht1.cell(row = x1, column = 4).value)
##    XfmrName = XfmrName.capitalize()
##    XFMR_set.add(XfmrName)
##   
##device_declaration = ''
##
##for n1 in XFMR_set:
##    Xfmr = str(n1 + '_XFMR')
##    device_declaration = str(device_declaration + Xfmr + ', ')
##
##end_string = ' : SIM_SUB_TRANSFORMER;'
##
##device_declaration = str(device_declaration[:-2] + end_string)
##
##
##Xfmr_sheet.cell(row = 1, column = 1, value = device_declaration)
##    
##XfmrRow = 2
##
##for n2 in XFMR_set:
##    Xfmr = n2 + '_XFMR'
##    
##    data1 = str(Xfmr + ".mInit(\n" + "\tSubstationName :=" + "\'" + str(n2) + "\',\n"    +"\tName:= 'Main XFMR',\n"    +  "\tVoltageLL:= 34500,\n"     + "\tDeviceDefinition:= NullTransformer,\n"  + 	"\tpFieldFirst_BI:= 0,\n"
##    +  "\tpFieldFirst_AI:= 0,\n"
##    +	"\tpFieldFirst_CNT:= 0,\n"
##    +  	"\tpFieldFirst_BO:= 0,\n"
##    +	"\tpEN:= 0);")
##
##    Xfmr_sheet.row_dimensions[XfmrRow].height = 150
##    Xfmr_sheet.cell(row = XfmrRow, column = 1, value = data1)
##    XfmrRow += 1
####    

wb.save(file_name)





    
    



    
            
            







